# In the database, there are 3 columns: 'Date of Birth', 'Full Name' and 'City'. We want to create another sheet for each city and to copy every row associated with that city
# to the new sheet, keeping the format equal to original. 

# for each row
    # check if city sheet already exists. If not, create sheet.
    # copy values from that row and paste in corresponding city sheet.

from openpyxl import load_workbook
from copy import copy

def crate_sheet(city, workbook_city, header_style):
    if city not in workbook_city.sheetnames:
        workbook_city.create_sheet(city)
        new_sheet = workbook_city[city]
        new_sheet["A1"].value = "Date of Birth"
        new_sheet["B1"].value = "Full Name"
        new_sheet["C1"].value = "City"
        new_sheet["A1"]._style = header_style
        new_sheet["B1"]._style = header_style
        new_sheet["C1"]._style = header_style
        for col_letter, dim in sheet_database.column_dimensions.items():
            if dim.width is not None:
                new_sheet.column_dimensions[col_letter].width = dim.width

def copy_data(sheet_from, sheet_to, row_from):
    row_to = sheet_to.max_row+1
    for j in range(1,4):
        cell_from = sheet_from.cell(row=row_from, column=j)
        cell_to = sheet_to.cell(row=row_to, column=j)
        cell_to.value = cell_from.value
        cell_to._style = copy(cell_from._style)

workbook_city = load_workbook("popdata.xlsx")
sheet_database = workbook_city['Sheet1']

last_line = sheet_database.max_row
header_style = copy(sheet_database["A1"]._style)

for i in range(2, last_line+1):
    city = sheet_database.cell(row=i,column=3).value
    if not city:
        break
    # create a sheet for each city
    crate_sheet(city, workbook_city, header_style)

    # transfer data to sheet
    sheet_to = workbook_city[city]
    copy_data(sheet_database, sheet_to, i)

workbook_city.save("popdata2.xlsx")
